import csv
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/Milltown-Archive/inventory.csv"
OUT = "/home/user/Milltown-Archive/inventory.xlsx"

with open(SRC, newline="", encoding="utf-8") as fh:
    rows = [r for r in csv.reader(fh) if any(c.strip() for c in r)]

header, data = rows[0], rows[1:]

wb = Workbook()
ws = wb.active
ws.title = "Inventory"

FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F4F4F")
body_font = Font(name=FONT, size=11)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(header)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 30

price_col = header.index("Price") + 1
date_col = header.index("Date Added") + 1
item_col = header.index("Item No.") + 1

for r in data:
    ws.append(r)

for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Price: stored as a number so it can be sorted and totalled,
    # displayed with the £ symbol. Blank stays blank.
    pc = ws.cell(row=row_idx, column=price_col)
    if isinstance(pc.value, str) and pc.value.strip():
        pc.value = float(pc.value.replace("£", "").strip())
    pc.number_format = '£#,##0.00'
    pc.alignment = Alignment(horizontal="right", vertical="top")

    # Item No.: the order the products were dictated in.
    ic = ws.cell(row=row_idx, column=item_col)
    if isinstance(ic.value, str) and ic.value.strip():
        ic.value = int(ic.value)
    ic.alignment = Alignment(horizontal="center", vertical="top")

    dc = ws.cell(row=row_idx, column=date_col)
    if isinstance(dc.value, str) and dc.value.strip():
        y, m, d = (int(p) for p in dc.value.split("-"))
        dc.value = date(y, m, d)
    dc.number_format = "DD/MM/YYYY"
    dc.alignment = Alignment(horizontal="left", vertical="top")

widths = {
    "Item No.": 9,
    "Product Name": 32,
    "Colour Clarity/Description": 26,
    "Defects": 34,
    "Size": 20,
    "Condition": 26,
    "SKU": 26,
    "Price": 11,
    "Date Added": 13,
}
for idx, name in enumerate(header, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = widths.get(name, 18)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

wb.save(OUT)
print(f"wrote {OUT} with {len(data)} product rows")
